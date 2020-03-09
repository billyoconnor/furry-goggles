#Script for handling data cleaning aspect of program
import os
import tkinter.messagebox
from tkinter import Tk
from tkinter.filedialog import askdirectory, askopenfilename
import numpy as np
import openpyxl
import pandas as pd
from pandas import read_excel

Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
tkinter.messagebox.showinfo(title="Upload", message="Please choose the workbook you wish to upload")# Popup window explaining about to ask for source workbook
path1 = askopenfilename() # show an "Open" dialog box and return the path to the selected file - Store path as variable 'path1'

wb1 = openpyxl.load_workbook(path1,data_only = True) # Create a variable 'wb1' to operate on spreadsheet in question
ws1 = wb1.worksheets[5] #Need to specify worksheet name

assName = os.path.basename(path1) # storing filename of workbook to using in Naming later
assName = os.path.splitext(assName)[0]
cutPoint = assName.find('Index') # Cutting filename to relevant bit
assNameLength = len(assName)
assName = assName[cutPoint:assNameLength] # Will give a strange value if there is no 'index' in the name

wb2 = openpyxl.Workbook()
ws2 = wb2.create_sheet(index=0, title='temp') # copying spreadsheet data to temporary spreadsheet in stripped format
for row in ws1:
    for cell in row:
        ws2[cell.coordinate].value = cell.value
ws2['B1'] = 'ExcelQ' #setting column names
ws2['C1'] = 'Value'

tkinter.messagebox.showinfo(title="Select Folder", message="Please choose a folder for the program to create temporary files")
folderPath = askdirectory(title='Select Folder') # shows dialog box and return the path
os.chdir(folderPath) #Change  working directory to output folder - maybe make a working folder
wb2.save('Output.xlsx') # Save temp file as 'output'

df1 = read_excel('ExcelExtract.xlsx') #defining table1 as prefilled spreadsheet
df2 = read_excel('Output.xlsx',0) # defining table2 as the data pulled from spreadsheet
df3 = pd.merge(df1,df2[['ExcelQ','Value']],on='ExcelQ', how='left') #Merging two tables on shared Excelq column (i.e performing a vlookup)
deter = np.where(df3['ExcelQ'] == 0, df3['Static Values'], df3['Value']) # Pulling across static values
deter = pd.DataFrame(deter).fillna(0)#replace NaN values with 0
df3['Value2'] = deter #adding a column which contains excel values and static values
df3.to_excel('dataframe.xlsx') #Saving output to excel for reading later

wb3 = openpyxl.load_workbook('dataframe.xlsx')
ws3 = wb3.worksheets[0]
ws3['F2'] = assName # saving a assignment name, technically redundent
wb3.save('dataframe.xlsx')
